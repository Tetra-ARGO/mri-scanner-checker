#!/usr/bin/python
import os
import shutil
import subprocess
from datetime import datetime
from openpyxl import load_workbook

def NIFTI_Creator():
    print("NIFTI_Creator script started...")

    rawFolderPath = '/mnt/argo/Studies/R2D3/Public/Analysis/raw'
    rsyncFolderPath = '/mnt/argo/Studies/R2D3/Public/Analysis/raw/raw_mri/WPC-9080'
    excelFilePath = '/mnt/argo/Studies/R2D3/Public/Analysis/misc/spreadsheets/niftTest.xlsx'

    # Ensure the folder paths end with a file separator
    if not rawFolderPath.endswith(os.sep):
        rawFolderPath += os.sep
    if not rsyncFolderPath.endswith(os.sep):
        rsyncFolderPath += os.sep

    # Ensure NIFTI_Converted directory exists
    niftiConvertedPath = os.path.join(rawFolderPath, 'NIFTI_Converted')
    os.makedirs(niftiConvertedPath, exist_ok=True)

    # Check if the paths exist
    if not os.path.exists(rsyncFolderPath):
        print(f"Error: rsync folder path '{rsyncFolderPath}' does not exist.")
        return
    if not os.path.exists(rawFolderPath):
        print(f"Error: raw folder path '{rawFolderPath}' does not exist.")
        return

    # Get a list of all items in the rsync folder
    rsync_dir = os.listdir(rsyncFolderPath)
    print(f"Contents of {rsyncFolderPath}: {rsync_dir}")

    # If rsync_dir is empty, exit
    if not rsync_dir:
        print(f"No items found in {rsyncFolderPath}. Exiting.")
        return

    # Extract directory names (scan dates)
    scanDates = [d for d in rsync_dir if os.path.isdir(os.path.join(rsyncFolderPath, d)) and d not in ['.', '..']]
    print(f"Extracted scanDates: {scanDates}")

    # If no scanDates are found, exit
    if not scanDates:
        print("No valid scan dates found. Exiting.")
        return

    # Open the Excel workbook
    try:
        workbook = load_workbook(excelFilePath)
        sheet = workbook.active
    except Exception as e:
        print(f"Error opening Excel file: {str(e)}")
        return

    # Loop through all items in the rsync folder
    print(f"Number of scan dates: {len(scanDates)}")
    for scanDate in scanDates:
        currentDir = os.path.join(rsyncFolderPath, scanDate)
        print(f"Entering directory: {currentDir}")

        if not os.path.exists(currentDir):
            print(f"Directory {currentDir} does not exist.")
            continue

        filesInDir = os.listdir(currentDir)
        subFolders = [f for f in filesInDir if os.path.isdir(os.path.join(currentDir, f))]

        print(f"Subfolders in {currentDir}: {subFolders}")

        for subDirName in subFolders:
            if subDirName in ['.', '..']:
                continue

            # Extract subject ID and scan ID
            parts = subDirName.split('_')
            subjID = parts[1] if len(parts) > 1 else parts[0]  # Extract SP000
            scanID = parts[2] if len(parts) > 2 else '1'  # If there's no suffix, assume scan ID is 1
            
            print(f"Current subjID is: {subjID}")
            print(f"Current scanID is: {scanID}")

            subjScanFolder = os.path.join(niftiConvertedPath, scanDate, f"{subjID}_{scanID}")

            print(f"Checking if NIFTI folder exists: {subjScanFolder}")
            if not os.path.isdir(subjScanFolder):
                try:
                    input_folder = os.path.join(currentDir, subDirName)
                    if not os.path.isdir(input_folder):
                        print(f"Error: Input folder does not exist: {input_folder}")
                        continue

                    print(f"Input folder for dcm2niix: {input_folder}")

                    dcm2niix_cmd = f'/home/hudlowe/anaconda3/pkgs/dcm2niix-1.0.20230411-h00ab1b0_0/bin/dcm2niix -f "%d" -z n -b n -s y "{input_folder}"'
                    print(f"Running dcm2niix: {dcm2niix_cmd}")
                    subprocess.run(dcm2niix_cmd, shell=True, check=True)

                    os.makedirs(subjScanFolder, exist_ok=True)
                    for file in os.listdir(input_folder):
                        if file.endswith(('.nii', '.bvec', '.bval')):
                            shutil.move(os.path.join(input_folder, file), os.path.join(subjScanFolder, file))

                    # Update the Excel file
                    all_vault_uids = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1) if row[0].value]
                    
                    vault_scan_id = '1' if scanID == '1' else '2'

                    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    new_row = [subjID, vault_scan_id, current_datetime, current_datetime]
                    sheet.append(new_row)
                    print(f"Added new row to Excel: {new_row}")

                except subprocess.CalledProcessError as e:
                    print(f"Error running dcm2niix: {str(e)}")
                except Exception as e:
                    print(f"Error processing folder {input_folder}: {str(e)}")
            else:
                print(f"The conversion has already been completed for subject ID: {subjID} and scan ID: {scanID}")

    # Save the Excel workbook
    try:
        workbook.save(excelFilePath)
        print(f"Excel file saved successfully at {excelFilePath}")
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")

    print("Conversion Completed!")

# Run the function
NIFTI_Creator()

